import time

from openpyxl import *
from datetime import datetime


def get_user_workbook():
    wb = load_workbook(filename="Quiz-System/Users_file.xlsx")
    sheet = wb['Sheet1']

    return sheet


def username_validation(user_input):
    sheet = get_user_workbook()

    for row in sheet.iter_rows(values_only=True):
        if user_input == str(row[2]):
            return True

    return False


def password_validation(user_pass):
    sheet = get_user_workbook()

    for row in sheet.iter_rows(values_only=True):
        if str(row[3]) == user_pass:
            return True

    return False


def print_user_data(user_name):
    sheet = get_user_workbook()
    quiz_list = list()

    for row in sheet.iter_rows(values_only=True):
        if row[2] == user_name:
            quiz_list.append(row)

    print(f'''Hello {user_name}, my records indicate that you have taken the following quizzes: (if empty none taken):\n
''')
    for i in quiz_list:
        if None not in i:
            print(f'''\t-{i[6]}, score: {i[7] * 100}%''')


def validate_available_quizzes():
    wb = load_workbook(filename="Quiz-System/Quizzes_files.xlsx")
    sheet = wb["Sheet1"]
    quiz_names = list()
    print("We currently have the following quizzes available:")

    i = 0
    for row in sheet.iter_rows(values_only=True):
        if i > 0:
            print(f'\t - Quiz Name: {row[0]}, # of Questions: {row[1]}, Estimated Duration: {row[2]} minutes')
            quiz_names.append(row[0])
        i += 1

    relevant_quiz = input("Which quiz would you like to register for: ")

    while relevant_quiz not in quiz_names:
        if relevant_quiz == "q":
            exit()
        relevant_quiz = input("Please enter a valid quiz or enter q to quit: ")

    return relevant_quiz


def start_quiz(quiz_name, user_name):
    print(f"The {quiz_name} quiz will begin in 5 seconds.")
    time.sleep(5)
    quiz_time = datetime.now()
    quiz_time = quiz_time.strftime("%d/%m/%Y %H:%M:%S")

    wb = load_workbook(f"Quiz-System/{quiz_name}/q_bank.xlsx")
    sheet = wb["Sheet1"]
    quiz_list = list()
    answer_list = list()
    possible_answers = ["True", "False", "T", "F"]

    for row in sheet.iter_rows(values_only=True, min_row=2, max_col=3):
        if None not in row[0:3]:
            quiz_list.append(row[0:3])
            quiz_answer = input(f"Q{row[0]}. {row[1]} (T/F): ")
            while quiz_answer not in possible_answers:
                quiz_answer = input(f"Please enter a valid answer: ")

            answer_list.append(quiz_answer)

    read_length = len(quiz_list)
    formatted_results_data = list()

    for i in range(read_length):
        temp_list = [user_name, quiz_name, quiz_time, quiz_list[i][1], quiz_list[i][2], bool(answer_list[i])]
        formatted_results_data.append(temp_list)

    return formatted_results_data


def add_quiz_results(formatted_results):
    wb = load_workbook(filename="Quiz-System/Results_file.xlsx")
    sheet = wb["Sheet1"]

    for i in formatted_results:
        sheet.append(i)
        wb.save(filename="Quiz-System/Results_file.xlsx")


def get_quiz_score():
    wb = load_workbook(filename="Quiz-System/Results_file.xlsx")
    sheet = wb["Sheet1"]

    correct_count = 0
    total_question_count = 0

    for row in sheet.iter_rows(values_only=True, min_row=2):
        if row[4] == row[5]:
            correct_count += 1

        total_question_count += 1

    grade_score = (correct_count / total_question_count) * 100
    return grade_score


def print_results(user_name, quiz_name):
    wb = load_workbook("Quiz-System/Results_file.xlsx")
    sheet = wb["Sheet1"]

    for row in sheet.iter_rows(values_only=True, min_row=2):
        if quiz_name == row[1] and user_name == row[0]:
            print(f"Q:{row[3]}, Your Answer: {row[4]}, Actual Answer: {row[5]}")


def clear_results():
    wb = load_workbook("Quiz-System/Results_file.xlsx")
    sheet = wb["Sheet1"]

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
            wb.save(filename="Quiz-System/Results_file.xlsx")


def store_user_info():
    print("In order to create an account please enter the following information : ")

    while True:
        first_name = input("First Name: ")
        last_name = input("Last Name: ")
        user_name = input("User Name: ")
        password = input("Password: ")
        age = input("age: ")
        faculty = input("Faculty: ")
        user_data = [first_name, last_name, user_name, str(password), str(age), faculty]

        wb = load_workbook("Quiz-System/Users_file.xlsx")
        sheet = wb["Sheet1"]
        unique_username = True

        for row in sheet.iter_rows(values_only=True, min_row=2):
            if user_name == row[2]:
                unique_username = False

        if not unique_username:
            print("The username is already in use, please enter a unique username")
            continue
        else:
            sheet.append(user_data)
            wb.save(filename="Quiz-System/Users_file.xlsx")
            print("You are signed up! You will now be redirected to login.")
            break


def add_taken_quiz_user_profile(active_user, relevant_quiz, quiz_score):
    wb = load_workbook(filename="Quiz-System/Users_file.xlsx")
    sheet = wb["Sheet1"]
    temp = None

    for row in sheet.iter_rows(values_only=True, min_row=2):
        if row[2] == active_user:
            temp = row

    temp_list = list(temp)
    temp_list.insert(6, relevant_quiz)
    temp_list.insert(7, quiz_score)
    sheet.append(temp_list)
    wb.save(filename="Quiz-System/Users_file.xlsx")













